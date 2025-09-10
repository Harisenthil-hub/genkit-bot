import os
import datetime
import openpyxl
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes

# ===============================
# Excel Setup
# ===============================
FILE_NAME = "attendance.xlsx"

def init_excel():
    if not os.path.exists(FILE_NAME):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name", "Date", "Check-in", "Check-out", "Work Done"])
        wb.save(FILE_NAME)

def log_checkin(user, time):
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([user, time.date(), time.strftime("%H:%M:%S"), "", ""])
    wb.save(FILE_NAME)

def log_checkout(user, time, work):
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb.active
    for row in reversed(list(ws.iter_rows(min_row=2))):
        if row[0].value == user and row[1].value == time.date() and not row[3].value:
            row[3].value = time.strftime("%H:%M:%S")
            row[4].value = work
            break
    wb.save(FILE_NAME)

def calculate_hours(checkin, checkout):
    try:
        in_time = datetime.datetime.strptime(checkin, "%H:%M:%S")
        out_time = datetime.datetime.strptime(checkout, "%H:%M:%S")
        duration = out_time - in_time
        hours = duration.seconds // 3600
        minutes = (duration.seconds % 3600) // 60
        return f"{hours}h {minutes}m"
    except:
        return "-"

def get_today_report(date):
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb.active
    report = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == date:
            report.append(row)
    return report

def get_weekly_report(end_date):
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb.active
    start_date = end_date - datetime.timedelta(days=6)
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, date, checkin, checkout, work = row
        if isinstance(date, datetime.date) and start_date <= date <= end_date:
            if name not in data:
                data[name] = {"days": 0, "minutes": 0}
            if checkin and checkout:
                in_time = datetime.datetime.strptime(checkin, "%H:%M:%S")
                out_time = datetime.datetime.strptime(checkout, "%H:%M:%S")
                minutes = (out_time - in_time).seconds // 60
                data[name]["days"] += 1
                data[name]["minutes"] += minutes
    return data

# ===============================
# Telegram Bot Handlers
# ===============================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("👋 Welcome to Genkit Attendance Bot!\nUse /help to see commands.")

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = (
        "📋 Available Commands:\n"
        "/checkin - Mark your check-in time\n"
        "/checkout <work> - Mark your check-out and work details\n"
        "/report - View today’s report\n"
        "/weeklyreport - View this week’s summary"
    )
    await update.message.reply_text(msg)

async def checkin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user.first_name
    now = datetime.datetime.now()
    log_checkin(user, now)
    await update.message.reply_text(f"✅ Check-in recorded at {now.strftime('%H:%M:%S')}")

async def checkout(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user.first_name
    now = datetime.datetime.now()
    work = " ".join(context.args) if context.args else "No work details provided"
    log_checkout(user, now, work)
    await update.message.reply_text(
        f"✅ Check-out recorded at {now.strftime('%H:%M:%S')}\n📝 Work: {work}"
    )

async def report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    today = datetime.date.today()
    records = get_today_report(today)
    if not records:
        await update.message.reply_text("No attendance records for today.")
        return
    msg = f"📅 Report for {today}:\n"
    for name, date, checkin, checkout, work in records:
        hours = calculate_hours(checkin, checkout) if checkin and checkout else "-"
        msg += f"{name} | In: {checkin} | Out: {checkout or '-'} | Hours: {hours} | Work: {work or '-'}\n"
    await update.message.reply_text(msg)

async def weeklyreport(update: Update, context: ContextTypes.DEFAULT_TYPE):
    today = datetime.date.today()
    data = get_weekly_report(today)
    if not data:
        await update.message.reply_text("No attendance records this week.")
        return
    msg = f"📅 Weekly Report ({today - datetime.timedelta(days=6)} → {today}):\n\n"
    for name, stats in data.items():
        total_hours = stats['minutes'] // 60
        total_minutes = stats['minutes'] % 60
        msg += f"{name} | Days: {stats['days']} | Total: {total_hours}h {total_minutes}m\n"
    await update.message.reply_text(msg)

# ===============================
# Main Function
# ===============================
def main():
    init_excel()
    token = os.getenv("BOT_TOKEN")
    if not token:
        print("❌ BOT_TOKEN not set in environment!")
        return

    app = Application.builder().token(token).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("checkin", checkin))
    app.add_handler(CommandHandler("checkout", checkout))
    app.add_handler(CommandHandler("report", report))
    app.add_handler(CommandHandler("weeklyreport", weeklyreport))

    print("✅ Bot started...")
    app.run_polling()

if __name__ == "__main__":
    main()
